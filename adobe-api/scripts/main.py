import logging
import os
import re
import sys
import time
from copy import copy
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook

from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
from adobe.pdfservices.operation.exception.exceptions import (
    ServiceApiException,
    ServiceUsageException,
    SdkException,
)
from adobe.pdfservices.operation.io.cloud_asset import CloudAsset
from adobe.pdfservices.operation.io.stream_asset import StreamAsset
from adobe.pdfservices.operation.pdf_services import PDFServices
from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
from adobe.pdfservices.operation.pdfjobs.jobs.export_pdf_job import ExportPDFJob
from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_params import ExportPDFParams
from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_target_format import ExportPDFTargetFormat
from adobe.pdfservices.operation.pdfjobs.result.export_pdf_result import ExportPDFResult


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

INPUT_DIR = Path("./inputs")
OUTPUT_DIR = Path("./outputs")
XLSX_DIR = OUTPUT_DIR / "xlsx"
COMBINED_WORKBOOK_PATH = OUTPUT_DIR / "combined.xlsx"

MAX_SHEET_TITLE_LEN = 31


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise EnvironmentError(
            f"Missing environment variable: {name}. "
            f"Set it before running this script."
        )
    return value


def load_credentials_from_json(path: Path) -> tuple[str, str]:
    import json

    if not path.exists():
        raise FileNotFoundError(f"Credentials JSON not found: {path}")

    payload = json.loads(path.read_text(encoding="utf-8"))
    client_id = payload.get("CLIENT_ID")
    client_secret = None
    secrets = payload.get("CLIENT_SECRETS")
    if isinstance(secrets, list) and secrets:
        client_secret = secrets[0]

    if not client_id or not client_secret:
        raise RuntimeError("Credentials JSON is missing CLIENT_ID or CLIENT_SECRETS")

    return client_id, client_secret


def sanitize_sheet_name(name: str, used_names: set) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = "Sheet"

    base = name[:MAX_SHEET_TITLE_LEN]
    candidate = base
    counter = 1

    while candidate in used_names:
        suffix = f"_{counter}"
        allowed = MAX_SHEET_TITLE_LEN - len(suffix)
        candidate = f"{base[:allowed]}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate


def get_pdf_services() -> PDFServices:
    client_id = os.getenv("PDF_SERVICES_CLIENT_ID")
    client_secret = os.getenv("PDF_SERVICES_CLIENT_SECRET")

    if not client_id or not client_secret:
        # Try to load from a JSON file path.
        env_creds_path = os.getenv("ADOBE_CREDENTIALS_JSON")
        if env_creds_path:
            creds_path = Path(env_creds_path)
        else:
            creds_path = Path(__file__).resolve().parents[1] / "761BlackAardwolf-4199879-OAuth Server-to-Server.json"

        if creds_path.exists():
            client_id, client_secret = load_credentials_from_json(creds_path)
        else:
            raise EnvironmentError(
                "Missing Adobe credentials. Set PDF_SERVICES_CLIENT_ID/PDF_SERVICES_CLIENT_SECRET "
                "or set ADOBE_CREDENTIALS_JSON to a valid credentials JSON path."
            )

    credentials = ServicePrincipalCredentials(
        client_id=client_id,
        client_secret=client_secret
    )
    return PDFServices(credentials=credentials)


def export_pdf_to_xlsx(pdf_services: PDFServices, pdf_path: Path, output_xlsx_path: Path) -> Path:
    logging.info("Exporting PDF to XLSX: %s", pdf_path.name)

    with open(pdf_path, "rb") as f:
        input_stream = f.read()

    input_asset = pdf_services.upload(
        input_stream=input_stream,
        mime_type=PDFServicesMediaType.PDF
    )

    export_pdf_params = ExportPDFParams(
        target_format=ExportPDFTargetFormat.XLSX
    )

    export_pdf_job = ExportPDFJob(
        input_asset=input_asset,
        export_pdf_params=export_pdf_params
    )

    location = pdf_services.submit(export_pdf_job)
    pdf_services_response = pdf_services.get_job_result(location, ExportPDFResult)

    result_asset: CloudAsset = pdf_services_response.get_result().get_asset()
    stream_asset: StreamAsset = pdf_services.get_content(result_asset)

    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_xlsx_path, "wb") as f:
        f.write(stream_asset.get_input_stream())

    logging.info("Saved XLSX: %s", output_xlsx_path)
    return output_xlsx_path


def copy_sheet_contents(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = source_ws.freeze_panes

    if source_ws.auto_filter and source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref


def merge_xlsx_files_to_workbook(xlsx_files: List[Path], combined_workbook_path: Path):
    logging.info("Merging %d XLSX files into %s", len(xlsx_files), combined_workbook_path)

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    used_sheet_names = set()

    for xlsx_file in xlsx_files:
        source_wb = load_workbook(xlsx_file)

        if not source_wb.sheetnames:
            logging.warning("No sheets found in %s; skipping", xlsx_file.name)
            continue

        if len(source_wb.sheetnames) == 1:
            source_ws = source_wb[source_wb.sheetnames[0]]
            sheet_title = sanitize_sheet_name(xlsx_file.stem, used_sheet_names)
            target_ws = master_wb.create_sheet(title=sheet_title)
            copy_sheet_contents(source_ws, target_ws)
        else:
            for source_sheet_name in source_wb.sheetnames:
                source_ws = source_wb[source_sheet_name]
                raw_name = f"{xlsx_file.stem}_{source_sheet_name}"
                sheet_title = sanitize_sheet_name(raw_name, used_sheet_names)
                target_ws = master_wb.create_sheet(title=sheet_title)
                copy_sheet_contents(source_ws, target_ws)

        source_wb.close()

    if not master_wb.sheetnames:
        master_wb.create_sheet(title="Empty")
        master_wb["Empty"]["A1"] = "No worksheets were merged."

    combined_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(combined_workbook_path)
    logging.info("Saved combined workbook: %s", combined_workbook_path)


def find_pdf_files(input_dir: Path) -> List[Path]:
    pdf_files = sorted([p for p in input_dir.glob("*.pdf") if p.is_file()])
    return pdf_files


def main():
    start = time.time()

    try:
        if not INPUT_DIR.exists():
            raise FileNotFoundError(
                f"Input directory not found: {INPUT_DIR.resolve()}\n"
                f"Create it and place your PDF files inside."
            )

        pdf_files = find_pdf_files(INPUT_DIR)
        if not pdf_files:
            raise FileNotFoundError(
                f"No PDF files found in {INPUT_DIR.resolve()}"
            )

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        XLSX_DIR.mkdir(parents=True, exist_ok=True)

        pdf_services = get_pdf_services()

        exported_xlsx_files = []
        failures = []

        for pdf_file in pdf_files:
            try:
                output_xlsx = XLSX_DIR / f"{pdf_file.stem}.xlsx"
                exported_file = export_pdf_to_xlsx(pdf_services, pdf_file, output_xlsx)
                exported_xlsx_files.append(exported_file)
            except (ServiceApiException, ServiceUsageException, SdkException) as e:
                logging.exception("Adobe export failed for %s: %s", pdf_file.name, e)
                failures.append((pdf_file.name, str(e)))
            except Exception as e:
                logging.exception("Unexpected failure for %s: %s", pdf_file.name, e)
                failures.append((pdf_file.name, str(e)))

        if not exported_xlsx_files:
            raise RuntimeError("No XLSX files were successfully exported.")

        merge_xlsx_files_to_workbook(exported_xlsx_files, COMBINED_WORKBOOK_PATH)

        elapsed = time.time() - start
        logging.info("Done in %.2f seconds", elapsed)

        if failures:
            logging.warning("Some files failed:")
            for fname, err in failures:
                logging.warning(" - %s | %s", fname, err)

    except Exception as e:
        logging.exception("Fatal error: %s", e)
        sys.exit(1)


if __name__ == "__main__":
    main()